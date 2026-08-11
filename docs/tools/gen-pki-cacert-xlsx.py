#!/usr/bin/env python3
# =============================================================================
# docs/PKI-CACERT-EXPORT.xlsx を生成する
# ---------------------------------------------------------------------------
# 自己証明書 (cacert.crt) の「配備 → 出力 (export) → イメージビルドへの受け渡し
# (build secret) → トラストストア取り込み」の一連の設計を Excel にまとめる。
#
# xlsx はバイナリのため差分レビューができない。内容の変更はこのスクリプトを直し、
# 再生成してコミットすること (単体で完結。引数不要)。
#
#   pip install openpyxl
#   python docs/tools/gen-pki-cacert-xlsx.py
#
# 対応する md:
#   docs/TLS-SELF-SIGNED-ALB.md      (3 章が export / build secret)
#   compose/pki/export/README.md
#   compose/pki/provided/README.md
# =============================================================================
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "PKI-CACERT-EXPORT.xlsx"

# --- 体裁 --------------------------------------------------------------------
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
HEAD_FONT = Font(bold=True, size=10, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
NOTE_FONT = Font(size=9, italic=True, color="595959")
MONO_FONT = Font(name="Consolas", size=9)
BASE_FONT = Font(size=10)
STAR_FILL = PatternFill("solid", fgColor="FFF2CC")   # ★重要行
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")    # 縞
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
WRAP_C = Alignment(vertical="top", horizontal="center", wrap_text=True)


def sheet(wb, name, title, note, headers, rows, widths, mono_cols=(), first=False):
    """1 シート = 見出し + 説明 + 表。star=True の行は強調する"""
    ws = wb.active if first else wb.create_sheet()
    ws.title = name

    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    ws["A2"] = note
    ws["A2"].font = NOTE_FONT
    ws["A2"].alignment = WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.row_dimensions[2].height = 30

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = WRAP_C
        cell.border = BORDER

    for r, row in enumerate(rows, start=5):
        star = bool(row and str(row[0]).startswith("★"))
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = MONO_FONT if c in mono_cols else BASE_FONT
            cell.alignment = WRAP
            cell.border = BORDER
            if star:
                cell.fill = STAR_FILL
            elif r % 2 == 0:
                cell.fill = ALT_FILL

    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    return ws


wb = Workbook()

# =============================================================================
# 1. 概要 — なぜ出力が要るのか
# =============================================================================
sheet(
    wb, "1.概要", "自己証明書 cacert.crt の配備・出力・取り込み — 全体像",
    "証明書の実体は named volume にあり実行時にしか見えない。一方 docker build が読めるのは "
    "ビルドコンテキスト (ホスト上のファイル) だけ。この 2 つを繋ぐために pki-init は "
    "cacert.crt をホストの compose/pki/export/ へも出力する。",
    ["#", "ステップ", "実体 / 場所", "内容", "コマンド"],
    [
        ["1", "投入 (任意)", "compose/pki/provided/cacert.crt",
         "受領済みの cacert.crt を置くと pki-init は CA を新規発行せず、その受領物を"
         "トラストアンカーとして配る (provided モード)。置かなければ自己署名 CA を自動発行 (generate)。",
         "cp /path/to/cacert.crt compose/pki/provided/"],
        ["2", "配備", "pki-init → named volume: pki",
         "cacert.crt と各サーバ証明書 (secure-api / ALB / MySQL) を発行し、全コンテナへ read-only で配る。",
         "docker compose up -d pki-init"],
        ["★3", "出力 (export)", "compose/pki/export/",
         "配備した cacert.crt をホストへ書き出す。docker build はここのファイルしか読めないため、"
         "ビルドへ渡すにはこの出力が要る。PEM 1 枚をそのまま出す (変換・リネームなし)。",
         "自動 (docker compose up -d pki-init)  /  ./pki-export.sh"],
        ["★4a", "ビルドへ受け渡し", "build secret (id=cacert)",
         "ベースイメージと同じ方式。ビルド中だけ tmpfs にマウントされ、イメージのレイヤーにも "
         "ビルドキャッシュにも残らない。JDK cacerts と jboss-truststore.p12 へ keytool で焼き込む。",
         "docker compose -f compose.yaml -f compose.build-secret.yaml build app-front app-back"],
        ["★4b", "受領物として固定", "compose/pki/provided/",
         "出力された cacert.crt をそのまま provided/ へ置くと、次回以降その CA を使い続ける。"
         "自動発行した CA を固定できる (入力と出力が往復する)。",
         "./pki-export.sh --to-provided"],
        ["5", "実行時取り込み", "entrypoint.sh (front / back)",
         "/mnt/pki/trust/*.crt を keytool で JDK 側 (cacerts のコピー) と JBoss 側 "
         "(jboss-truststore.p12) の 2 ストアへ取り込む。従来からの経路でそのまま残る。",
         "docker compose logs app-front | grep truststore"],
        ["6", "検証", "tls-verifier / tls-probe",
         "secure-api・ALB への HTTPS 呼び出しが JDK 経路 / JBoss 経路の両方で通ること、"
         "および取り込まれた証明書の SHA-256 が出力物と一致することを確認する。",
         "./verify-tls.sh"],
    ],
    [6, 18, 30, 62, 60], mono_cols=(5,), first=True,
)

# =============================================================================
# 2. 出力ファイル一覧
# =============================================================================
sheet(
    wb, "2.出力ファイル", "compose/pki/export/ に出力されるファイル",
    "出力は毎回まるごと作り直す。特に cacert.key は provided モードの分岐そのものを変えるため、"
    "古い鍵が残って証明書と食い違う (pki-init の起動失敗) 事故を防ぐ。",
    ["ファイル", "出力条件", "内容", "主な使い道"],
    [
        ["★cacert.crt", "常に",
         "唯一のトラストアンカー。受領物を使っている場合は受領物のバイト列そのもの (PEM 1 枚)。",
         "build secret の入力 / compose/pki/provided/ へ置いて CA を固定"],
        ["cacert.key", "秘密鍵がある場合のみ\n(PKI_EXPORT_KEY=1)",
         "CA の秘密鍵。generate モード、または受領物に鍵が同梱されている場合に存在する。",
         "provided/ へ cacert.crt と一緒に置くとパターン A (受領 CA が全証明書を発行) を維持できる"],
        ["verify-bundle.crt", "常に",
         "このスタックのサーバ証明書を検証できる CA の集合。= cacert.crt (+ 鍵なし時のみ local-test-ca.crt)。",
         "curl --cacert などホストからの検証"],
        ["trust/cacert.crt", "常に",
         "front/back のトラストストアへ入る本命。cacert.crt と同一。ファイル名が keytool の alias になる。",
         "複数証明書をまとめて焼き込むビルドの入力"],
        ["trust/alb-selfsigned.crt", "常に",
         "ALB に自己署名リーフを適用したとき用 (その 1 枚だけを信頼するパターン B)。",
         "同上"],
        ["trust/local-test-ca.crt", "鍵なし受領 かつ\nPKI_TRUST_LOCAL_CA=1",
         "受領 CA の鍵が無いときにサーバ証明書を発行するローカル CA。",
         "同上 (これを入れないと HTTPS 疎通の正常系が通らない)"],
        ["MANIFEST.txt", "常に",
         "出力日時 / モード / cacert.crt の subject・issuer・有効期限・SHA-256 指紋 / 各ファイルの sha256 / 使い方。",
         "ビルドへ渡した証明書と、取り込まれた証明書の同一性確認"],
    ],
    [26, 22, 62, 62],
)

# =============================================================================
# 3. build secret による取り込み
# =============================================================================
sheet(
    wb, "3.build secret", "イメージビルドへの受け渡し (BuildKit の build secret)",
    "社内標準ベースイメージが行っているのと同じ方式。docker/front/Dockerfile と "
    "docker/back/Dockerfile にも同一の取り込みを実装済みなので、ベースイメージを差し替えなくても"
    "ローカルで同じ経路を検証できる。secret を渡さなければ何もしないため既存のビルド手順は壊れない。",
    ["項目", "内容", "備考"],
    [
        ["★渡し方 (compose)",
         "docker compose -f compose.yaml -f compose.build-secret.yaml build app-front app-back",
         "compose.build-secret.yaml が build.secrets: [cacert] と secrets.cacert.file を配線する"],
        ["★渡し方 (docker build)",
         "docker build --secret id=cacert,src=compose/pki/export/cacert.crt -f front/Dockerfile ./docker",
         "ベースイメージのビルドはこの形。BuildKit が必要 (DOCKER_BUILDKIT=1)"],
        ["Dockerfile 側",
         "RUN --mount=type=secret,id=cacert,target=/run/secrets/cacert.crt ...",
         "ビルド中だけ tmpfs にマウントされ、イメージのレイヤーにもビルドキャッシュにも残らない"],
        ["取り込み先 1 (JVM)",
         'keytool -importcert -alias cacert -file /run/secrets/cacert.crt -keystore "${JAVA_HOME}/lib/security/cacerts"',
         "ビルド時は root なので JDK 同梱 cacerts を直接更新できる (実行時は UID 185 のため不可)"],
        ["取り込み先 2 (JBoss EAP)",
         'keytool -importcert -alias cacert -file /run/secrets/cacert.crt '
         '-keystore "${JBOSS_HOME}/standalone/configuration/jboss-truststore.p12" -storetype PKCS12',
         "cli/elytron-truststore.cli の key-store=appTrustStore が参照するファイルそのもの"],
        ["取り込み記録",
         "/opt/pki/build-import.txt",
         "日時 / alias / 取り込み先 / 証明書の subject・issuer・SHA-256。entrypoint が起動時にログへ出す"],
        ["secret 未指定のとき",
         "何もしない (ログに [build][cacert] ... スキップします)",
         "従来どおり entrypoint.sh が起動時に ${PKI_TRUST_DIR} から取り込む"],
        ["確認方法",
         "docker compose logs app-front | grep 'truststore[build]'\n"
         "docker compose exec app-front cat /opt/pki/build-import.txt\n"
         "curl -s http://localhost:8080/tls-probe/truststore | jq -r '.stores[].cacertSha256'",
         "最後の指紋が compose/pki/export/cacert.crt の SHA-256 と一致すれば同一の証明書"],
        ["build arg",
         "CACERT_ALIAS=cacert / JVM_TRUSTSTORE_PASSWORD=changeit / JBOSS_TRUSTSTORE_PASSWORD=changeit",
         "既定値のまま使う想定 (テスト環境専用)"],
    ],
    [26, 76, 62], mono_cols=(2,),
)

# =============================================================================
# 4. ビルド時 vs 実行時
# =============================================================================
sheet(
    wb, "4.ビルド時と実行時", "ビルド時取り込み (build secret) と 実行時取り込み (entrypoint) の比較",
    "両者は併用できる。PKI_TRUST_DIR に *.crt が無ければ実行時取り込みはスキップされ、"
    "イメージへ焼き込んだトラストストアがそのまま使われる (= pki volume 無しでも HTTPS が通る)。",
    ["観点", "ビルド時取り込み (build secret)", "実行時取り込み (entrypoint)"],
    [
        ["実行ユーザ", "root — JDK 同梱 cacerts を直接更新できる",
         "jboss (UID 185) — 書き換えられないので /tmp/pki/cacerts へコピーして追加する"],
        ["JDK 側の実体", "${JAVA_HOME}/lib/security/cacerts (イメージ内)",
         "/tmp/pki/cacerts (-Djavax.net.ssl.trustStore で JVM へ指定)"],
        ["JBoss 側の実体", "${JBOSS_HOME}/standalone/configuration/jboss-truststore.p12 をビルド時に作り込む",
         "同じパスを毎起動で削除して作り直す"],
        ["証明書の入手元", "build secret (ホスト上のファイル = compose/pki/export/cacert.crt)",
         "${PKI_TRUST_DIR} (既定 /mnt/pki/trust。named volume 経由)"],
        ["★証明書を差し替えたら", "イメージの再ビルドが必要",
         "コンテナの再起動だけで反映される"],
        ["pki volume 無しで動くか", "動く (イメージに焼き込み済み)",
         "動かない (取り込みがスキップされる)"],
        ["取り込む証明書の数", "cacert.crt 1 枚 (secret で渡した分)",
         "${PKI_TRUST_DIR}/*.crt すべて (cacert / alb-selfsigned / local-test-ca)"],
        ["ログ", "truststore[build]: (entrypoint が /opt/pki/build-import.txt を出力)",
         "truststore[jdk]: / truststore[jboss]:"],
        ["実 AWS での使いどころ", "CI でイメージへ焼き込む。読み取り専用ルートFSでも確実",
         "S3 / Secrets Manager から取得して配置する処理に置き換える"],
    ],
    [24, 62, 62],
)

# =============================================================================
# 5. 環境変数
# =============================================================================
sheet(
    wb, "5.環境変数", "関連する環境変数 (.env / compose.yaml / build arg)",
    "PKI_EXPORT_DIR と PKI_PROVIDED_DIR は、compose.yaml の environment (コンテナ内パス。固定値) と "
    "volumes の ${...} (ホストパス。.env で解決) で別スコープになっている点に注意。",
    ["変数", "設定場所", "既定値", "説明"],
    [
        ["★PKI_EXPORT_DIR", ".env (ホストパス)", "./compose/pki/export",
         "cacert.crt 一式の出力先。pki-init へ /export として read-write で bind mount する。"
         "ベースイメージのビルドコンテキスト配下を直接指定してもよい (例: ../base-image/secrets)。"
         "マウントが無い場合は出力せず警告のみ"],
        ["★PKI_EXPORT_KEY", ".env", "1",
         "CA 秘密鍵 (cacert.key) も出力するか。1 = 出力する。0 にすると鍵を持ち出さない "
         "(その場合、出力物を provided/ へ置くと必ずパターン B になる)"],
        ["★PKI_BUILD_SECRET_FILE", ".env", "./compose/pki/export/cacert.crt",
         "compose.build-secret.yaml が build secret としてビルドへ渡すファイル。"
         "受領物を直接渡すなら ./compose/pki/provided/cacert.crt を指定する"],
        ["PKI_PROVIDED_DIR", ".env (ホストパス)", "./compose/pki/provided",
         "受領した cacert.crt の投入口。pki-init へ /provided として read-only で bind mount する"],
        ["PKI_MODE", ".env", "auto",
         "auto = cacert.crt があれば provided / provided = 必須 (無ければ起動失敗) / generate = 受領物を無視して自動発行"],
        ["PKI_TRUST_LOCAL_CA", ".env", "1",
         "鍵なし受領時に local-test-ca を front/back のトラストストアへ入れるか"],
        ["PKI_TRUST_DIR", "compose.yaml (front/back)", "/mnt/pki/trust",
         "entrypoint が実行時に取り込む *.crt の置き場。ここに *.crt が無ければ実行時取り込みはスキップされる"],
        ["JBOSS_TRUSTSTORE_FILE", "compose.yaml (front/back)",
         "/opt/server/standalone/configuration/jboss-truststore.p12",
         "JBoss (Elytron) 側トラストストア。elytron-truststore.cli の path と一致必須"],
        ["CACERT_ALIAS", "build arg (Dockerfile)", "cacert",
         "ビルド時取り込みで使う keytool の alias"],
        ["JVM_TRUSTSTORE_PASSWORD", "build arg / 環境変数", "changeit",
         "JDK 側トラストストアのパスワード (JDK 既定値)"],
        ["JBOSS_TRUSTSTORE_PASSWORD", "build arg / 環境変数", "changeit",
         "JBoss 側トラストストアのパスワード"],
        ["CACERT_BUILD_MARKER", "環境変数 (front/back)", "/opt/pki/build-import.txt",
         "ビルド時取り込みの記録。entrypoint がこれを読んで truststore[build]: としてログに出す"],
    ],
    [30, 26, 32, 74],
)

# =============================================================================
# 6. 手順書
# =============================================================================
sheet(
    wb, "6.手順", "作業手順 (よく使う 4 パターン)",
    "いずれも AWS には接続しない。EAP_BASE_IMAGE だけは .env での設定が必要。",
    ["ケース", "#", "コマンド", "確認 / 期待値"],
    [
        ["★A. 自動発行した CA をビルドへ焼き込む", "1",
         "docker compose up -d pki-init",
         "compose/pki/export/cacert.crt が出力される"],
        ["", "2", "ls compose/pki/export/",
         "cacert.crt / cacert.key / verify-bundle.crt / trust/ / MANIFEST.txt"],
        ["", "3",
         "docker compose -f compose.yaml -f compose.build-secret.yaml build app-front app-back",
         "ビルドログに [build][cacert] cacert.crt を ... 取り込みました"],
        ["", "4",
         "docker compose -f compose.yaml -f compose.build-secret.yaml up -d",
         "起動ログに truststore[build]:"],
        ["", "5", "./verify-tls.sh", "コンテナ内検証がすべて PASS"],
        ["★B. 出力した CA を受領物として固定する", "1", "./pki-export.sh --to-provided",
         "compose/pki/provided/ に cacert.crt (+ cacert.key) が置かれる"],
        ["", "2", "docker compose restart pki-init",
         "ログが MODE: provided になる"],
        ["", "3",
         "docker compose restart secure-api alb mysql app-front app-back",
         "指紋が変わらないこと (docker compose logs pki-init | grep SHA-256)"],
        ["C. 受領した cacert.crt を使う", "1",
         "cp /path/to/受領した/cacert.crt compose/pki/provided/cacert.crt",
         "鍵があれば cacert.key も一緒に置く"],
        ["", "2", "docker compose up -d --build", "docker compose logs pki-init | grep 'MODE:' が provided"],
        ["", "3",
         "docker compose -f compose.yaml -f compose.build-secret.yaml build app-front app-back",
         "受領物がそのままイメージへ焼き込まれる (export 経由)"],
        ["D. ベースイメージのビルドへ渡す", "1",
         "echo 'PKI_EXPORT_DIR=../base-image/secrets' >> .env",
         "出力先をビルドコンテキスト上の所定ディレクトリへ向ける"],
        ["", "2", "docker compose up -d pki-init", "../base-image/secrets/cacert.crt が出力される"],
        ["", "3",
         "docker build --secret id=cacert,src=secrets/cacert.crt -t base-image .",
         "(ベースイメージ側のリポジトリで実行。./pki-export.sh --to <dir> でも配置できる)"],
    ],
    [36, 5, 74, 56], mono_cols=(3,),
)

# =============================================================================
# 7. トラブルシューティング
# =============================================================================
sheet(
    wb, "7.トラブル対応", "トラブルシューティング (出力 / build secret 関連)",
    "詳細は docs/TLS-SELF-SIGNED-ALB.md の 10 章 (トラブルシューティング) も参照。",
    ["症状", "原因", "対処"],
    [
        ["compose/pki/export/ に何も出力されない",
         "export の bind mount が無い (pki-init のログに『/export が無いため出力しません』)",
         "compose.yaml の pki-init に ${PKI_EXPORT_DIR:-./compose/pki/export}:/export があるか確認。"
         "docker compose up -d pki-init で作り直すか ./pki-export.sh で取り出す"],
        ["WARN: export: /export へ書き込めません",
         "出力先を :ro でマウントしている / ホスト側の書き込み権限が無い",
         "マウント指定から :ro を外す。PKI_EXPORT_DIR に指定したパスの権限を確認する"],
        ["docker compose build が secret ... no such file or directory で失敗",
         "compose/pki/export/cacert.crt がまだ出力されていない",
         "先に docker compose up -d pki-init (もしくは ./pki-export.sh)。"
         "このエラーは -f compose.build-secret.yaml を付けたときだけ起きる"],
        ["ビルドログに [build][cacert] ... スキップします",
         "secret がビルドへ届いていない",
         "-f compose.build-secret.yaml を付けているか / docker build なら --secret id=cacert,src=... の id / "
         "BuildKit が有効か (DOCKER_BUILDKIT=1) を確認"],
        ["★出力物を provided/ へ置いたら pki-init が "
         "『cacert.key は cacert.crt の秘密鍵ではありません』で落ちる",
         "配置先に古い cacert.key が残っていて、証明書と対になっていない",
         "compose/pki/provided/cacert.key を消してから置き直す "
         "(./pki-export.sh --to-provided は自動で消す)"],
        ["ビルド時に焼き込んだのに実行時の中身が違う",
         "PKI_TRUST_DIR に *.crt があると JBoss 側ストアは毎起動で作り直される "
         "(JDK 側は cacerts のコピーなのでビルド時の分も残る)",
         "docker compose logs app-front | grep truststore で [build] と [jboss] の両方を確認する"],
        ["tls-probe の cacertSha256 が出力物と一致しない",
         "別の CA が入っている / PKI を作り直した後に front/back を再起動していない",
         "docker compose restart app-front app-back。"
         "それでも直らなければ docker compose logs pki-init | grep SHA-256 と突き合わせる"],
        ["出力物を provided/ へ置いたら local-test-ca が作られた",
         "cacert.key を出力していない (PKI_EXPORT_KEY=0) / 一緒に置いていない",
         "鍵も置けばパターン A になる。鍵を出したくない場合はパターン B が正しい挙動"],
    ],
    [46, 52, 66],
)

# =============================================================================
# 8. 実 AWS への読み替え
# =============================================================================
sheet(
    wb, "8.実AWSへの読み替え", "ローカル構成と実 AWS の対応",
    "ローカルは AWS に一切接続しない。証明書の『入手経路』だけが変わり、"
    "イメージへの取り込み方 (build secret + keytool) はそのまま通用する。",
    ["ローカル", "実 AWS", "備考"],
    [
        ["compose/pki/provided/cacert.crt (受領物)",
         "社内 CA / 取引先から連携される自己署名ルート証明書そのもの",
         "ファイル名・形式ともそのまま"],
        ["★compose/pki/export/cacert.crt (出力)",
         "CI がビルド前に取得する CA 証明書 "
         "(S3 / Secrets Manager / SSM Parameter Store / 社内配布サーバ)",
         "取得してビルドコンテキストへ置く処理に置き換える"],
        ["★--secret id=cacert でのビルド時取り込み",
         "同じ。CodeBuild なら Secrets Manager から取得したファイルを docker build --secret に渡す",
         "イメージ層に残さないという性質もそのまま"],
        ["pki-init (openssl)", "AWS Private CA (ACM PCA) / 社内 CA",
         "ローカルは発行も兼ねるが、実 AWS では発行は CA 側の責務"],
        ["local-test-ca", "対応物なし",
         "受領 CA の鍵を持たないローカル事情のための代替。実 AWS では発生しない"],
        ["named volume pki", "EFS / サイドカーで取得したファイル / イメージ同梱",
         "読み取り専用ルートFSなら イメージ同梱 (= build secret 方式) が確実"],
        ["entrypoint の keytool 取り込み",
         "同じ。ECS では証明書を S3 / Secrets Manager / SSM から取得して PKI_TRUST_DIR へ配置する処理に置き換える",
         "-Djavax.net.ssl.trustStore が指すファイルは書き込み可能なパス (/tmp 等) に置く"],
        ["secure-api (WireMock HTTPS)", "自己証明書で HTTPS を要求する外部 API / 社内システム", "—"],
    ],
    [46, 62, 56],
)

wb.save(OUT)
print("wrote %s" % OUT)
